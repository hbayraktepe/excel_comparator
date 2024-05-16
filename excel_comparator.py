import sys
import os
import pandas as pd
from PyQt5 import QtWidgets, QtGui
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import traceback

class ExcelComparator(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.file_path1 = None
        self.file_path2 = None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Excel Comparator")
        self.setGeometry(100, 100, 600, 400)

        self.setWindowIcon(QtGui.QIcon(self.resource_path("x.jpg")))

        central_widget = QtWidgets.QWidget()
        self.setCentralWidget(central_widget)

        layout = QtWidgets.QVBoxLayout()

        # File selection labels
        self.label1 = QtWidgets.QLabel("Select Current Excel File")
        layout.addWidget(self.label1)
        self.button1 = QtWidgets.QPushButton("Browse")
        self.button1.clicked.connect(lambda: self.load_file(1))
        layout.addWidget(self.button1)

        self.label2 = QtWidgets.QLabel("Select Previous Excel File")
        layout.addWidget(self.label2)
        self.button2 = QtWidgets.QPushButton("Browse")
        self.button2.clicked.connect(lambda: self.load_file(2))
        layout.addWidget(self.button2)

        # Compare button
        self.compare_button = QtWidgets.QPushButton("Compare and Export to Excel")
        self.compare_button.clicked.connect(self.compare_and_export)
        layout.addWidget(self.compare_button)

        central_widget.setLayout(layout)

        self.setStyleSheet(self.load_stylesheet("style.qss"))

    def load_file(self, file_number):
        filepath, _ = QFileDialog.getOpenFileName(self, "Select File", "", "Excel files (*.xlsx *.xls)")
        if filepath:
            if file_number == 1:
                self.label1.setText(filepath)
                self.file_path1 = filepath
            elif file_number == 2:
                self.label2.setText(filepath)
                self.file_path2 = filepath

    def compare_and_export(self):
        file_path1 = self.file_path1
        file_path2 = self.file_path2

        if not file_path1 or not file_path2:
            QMessageBox.warning(self, "Warning", "Please select both files.")
            return

        try:
            df_current = pd.read_excel(file_path1)
            df_previous = pd.read_excel(file_path2)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error loading files: {e}")
            self.log_error(e)
            return

        # Drop the 'Order Column No.' column for comparison
        df_current.drop(columns=['Sip. Klm. No.'], inplace=True)
        df_previous.drop(columns=['Sip. Klm. No.'], inplace=True)

        if 'Material Description' not in df_current.columns or 'Material No' not in df_current.columns or \
           'Material Description' not in df_previous.columns or 'Material No' not in df_previous.columns:
            QMessageBox.critical(self, "Error", "'Material Description' and 'Material No' columns must be present in both files.")
            return

        try:
            new_entries, deleted_entries, changed_entries = self.compare_files(df_current, df_previous)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error comparing files: {e}")
            self.log_error(e)
            return

        output_path = 'Comparison_Result.xlsx'

        try:
            with ExcelWriter(output_path, engine='openpyxl') as writer:
                new_entries.to_excel(writer, sheet_name='New Entries', index=False)
                deleted_entries.to_excel(writer, sheet_name='Deleted Entries', index=False)
                changed_entries.to_excel(writer, sheet_name='Changed Entries', index=False)
            QMessageBox.information(self, "Success", f"Results have been written to {output_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error writing results: {e}")
            self.log_error(e)
            return

        try:
            self.highlight_differences(output_path)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error highlighting differences: {e}")
            self.log_error(e)

    @staticmethod
    def compare_files(df_current, df_previous):
        # Sort both dataframes by 'Material No' and 'Material Description'
        df_current_sorted = df_current.sort_values(by=['Material No', 'Material Description'])
        df_previous_sorted = df_previous.sort_values(by=['Material No', 'Material Description'])

        # Reset indices
        df_current_sorted.reset_index(drop=True, inplace=True)
        df_previous_sorted.reset_index(drop=True, inplace=True)

        new_entries = pd.DataFrame()
        deleted_entries = pd.DataFrame()
        changed_entries = pd.DataFrame()

        grouped_current = df_current_sorted.groupby(['Material No', 'Material Description'])
        grouped_previous = df_previous_sorted.groupby(['Material No', 'Material Description'])

        for group_name, group in grouped_current:
            if group_name not in grouped_previous.groups:
                new_entries = pd.concat([new_entries, group])
            else:
                previous_group = grouped_previous.get_group(group_name)
                current_dates = set(group['Delivery Date'])
                previous_dates = set(previous_group['Delivery Date'])
                new_dates = current_dates - previous_dates
                deleted_dates = previous_dates - current_dates
                common_dates = current_dates & previous_dates

                new_entries = pd.concat([new_entries, group[group['Delivery Date'].isin(new_dates)]])
                deleted_entries = pd.concat([deleted_entries, previous_group[previous_group['Delivery Date'].isin(deleted_dates)]])
                for date in common_dates:
                    current_rows = group[group['Delivery Date'] == date]
                    previous_rows = previous_group[previous_group['Delivery Date'] == date]
                    for _, current_row in current_rows.iterrows():
                        for _, previous_row in previous_rows.iterrows():
                            if current_row['Quantity'] != previous_row['Quantity']:
                                changed_row = current_row.copy()
                                changed_row['Previous Quantity'] = previous_row['Quantity']
                                changed_entries = pd.concat([changed_entries, changed_row.to_frame().T])

        for group_name, group in grouped_previous:
            if group_name not in grouped_current.groups:
                deleted_entries = pd.concat([deleted_entries, group])

        return new_entries, deleted_entries, changed_entries

    @staticmethod
    def highlight_differences(output_path):
        wb = load_workbook(output_path)
        ws = wb['Changed Entries']
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        df = pd.read_excel(output_path, sheet_name='Changed Entries')
        row_offset = 2  # Data starts from the second row (to skip the header)

        for index, row in df.iterrows():
            for col in df.columns:
                if col == 'Quantity' or col == 'Previous Quantity':
                    excel_row = index + row_offset
                    excel_col = df.columns.get_loc(col) + 1
                    cell = ws.cell(row=excel_row, column=excel_col)
                    cell.fill = fill

        wb.save(output_path)

    @staticmethod
    def log_error(e):
        with open("error_log.txt", "a") as f:
            f.write(f"{traceback.format_exc()}\n")

    def load_stylesheet(self, filename):
        path = self.resource_path(filename)
        with open(path, "r") as file:
            return file.read()

    @staticmethod
    def resource_path(relative_path):
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")

        return os.path.join(base_path, relative_path)


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    comparator = ExcelComparator()
    comparator.show()
    sys.exit(app.exec_())
